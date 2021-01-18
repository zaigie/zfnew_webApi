import datetime
import os
import time
import traceback

import json
import requests
import openpyxl
from bs4 import BeautifulSoup
from api import GetInfo, Login, PLogin, Personal, Infos, Search
from django.utils.encoding import escape_uri_path
from django.http import HttpResponse, JsonResponse, FileResponse
from info.models import Students, Teachers
from mp.models import Config
from openpyxl.styles import Font, colors, Alignment

with open('config.json', mode='r', encoding='utf-8') as f:
    config = json.loads(f.read())
base_url = config["base_url"]


def index(request):
    return HttpResponse('info_index here')

def calSex(id):
    sexNum = id[16:17]
    if int(sexNum)%2==0:
        return 2
    else:
        return 1

def diffList(list1,list2):
    return [x for x in list1 if x not in list2]

def mywarn(text,desp,xh,pswd):
    ServerChan = config["ServerChan"]
    text = text
    errData = {'err':text+'，请返回重试'} if "错误" in text else {'err':text+'，建议访问一下“课程通知”以便刷新cookies'}
    if ServerChan == "none":
        return HttpResponse(json.dumps(errData, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    else:
        requests.get(ServerChan + 'text=' + text + '&desp=' + desp + '\n' + str(xh) + '\n' + str(pswd))
        return HttpResponse(json.dumps(errData, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def cacheData(xh, filename):
    docurl = 'data/' + str(xh)[0:2] + '/' + str(xh) + '/'
    fileurl = docurl + str(filename) + '.json'
    if not os.path.exists(docurl):
        os.makedirs(docurl)
    else:
        if not os.path.exists(fileurl):
            return
        else:
            with open(fileurl, mode='r', encoding='utf-8') as o:
                result = json.loads(o.read())
                if result.get("err"):
                    return
                return result


def newData(xh, filename, content):
    docurl = 'data/' + str(xh)[0:2] + '/' + str(xh) + '/'
    fileurl = docurl + str(filename) + '.json'
    if not os.path.exists(docurl):
        os.makedirs(docurl)
        with open(fileurl, mode='w', encoding='utf-8') as n:
            n.write(content)
    else:
        with open(fileurl, mode='w', encoding='utf-8') as n:
            n.write(content)
    # if not os.path.exists(fileurl):
    #     with open(fileurl, mode='w', encoding='utf-8') as n:
    #         n.write(content)


def writeLog(content):
    date = datetime.datetime.now().strftime('%Y-%m-%d')
    filename = 'mylogs/' + date + '.log'
    if not os.path.exists(filename):
        with open(filename, mode='w', encoding='utf-8') as n:
            n.write('【%s】的日志记录' % date)
    with open(filename, mode='a', encoding='utf-8') as l:
        l.write('\n%s' % content)


def update_cookies(xh, pswd):
    try:
        stu = Students.objects.get(studentId=int(xh))
        refreshTimes = int(stu.refreshTimes)
        startTime = time.time()
        content = ('【%s】[%s]更新cookies' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
        writeLog(content)
        # print('原cookies：')
        # print('{JSESSIONID:%s,route:%s}' % (stu.JSESSIONID,stu.route))
        lgn = Login(base_url=base_url)
        lgn.login(xh, pswd)
        if lgn.runcode == 1:
            cookies = lgn.cookies
            # person = GetInfo(base_url=base_url, cookies=cookies)
            NJSESSIONID = requests.utils.dict_from_cookiejar(cookies)["JSESSIONID"]
            nroute = requests.utils.dict_from_cookiejar(cookies)["route"]
            updateTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            refreshTimes += 1
            Students.objects.filter(studentId=int(xh)).update(JSESSIONID=NJSESSIONID, route=nroute,
                                                              refreshTimes=refreshTimes, updateTime=updateTime)
            endTime = time.time()
            spendTime = endTime - startTime
            # print('新cookies:')
            content = ('【%s】更新cookies成功，耗时%.2fs' % (datetime.datetime.now().strftime('%H:%M:%S'), spendTime))
            writeLog(content)
            person = GetInfo(base_url=base_url, cookies=cookies)
            pinfo = person.get_pinfo()
            if stu.email == "无":
                Students.objects.filter(studentId=int(xh)).update(email=pinfo["email"])
            # print(pinfo)
            filename = ('Pinfo')
            newData(xh, filename, json.dumps(pinfo, ensure_ascii=False))
            # print(requests.utils.dict_from_cookiejar(cookies))
            return cookies
        else:
            content = ('【%s】[%s]更新cookies时网络或其他错误！' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'网络或token问题，请返回重试'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
    except Exception as e:
        if str(e) == "'NoneType' object has no attribute 'get'":
            return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
            return update_cookies(xh, pswd)
        else:
            traceback.print_exc()
            return mywarn("更新cookies未知错误",str(e),xh,pswd)

def writeToExcel(json,saveUrl):
    lastCourses = json["lastCourses"]
    res = json["res"]
    excel = openpyxl.Workbook()
    sheet1 = excel.create_sheet('sheet1', index=0)
    sheet1.cell(row=1,column=1,value="学号").alignment = Alignment(horizontal='center', vertical='center')
    sheet1.cell(row=1,column=2,value="姓名").alignment = Alignment(horizontal='center', vertical='center')
    sheet1.column_dimensions['A'].width = 15
    for c in range(0,len(lastCourses)):
        sheet1.cell(row=1, column=c + 3, value=lastCourses[c]).alignment = Alignment(horizontal='center', vertical='center')
        # sheet1.column_dimensions[chr(67+c)].width = 8
    for items in range(0,len(res)):
        sheet1.cell(row=items+2,column=1,value=res[items]["xh"]).alignment = Alignment(horizontal='center', vertical='center')
        sheet1.cell(row=items+2,column=2,value=res[items]["name"]).alignment = Alignment(horizontal='center', vertical='center')
        for n in range(0,len(res[items]["grades"])):
            for cs in range(0,len(lastCourses)):
                if res[items]["grades"][n]["n"] == lastCourses[cs]:
                    try:
                        sheet1.cell(row=items+2,column=cs+3,value=int(res[items]["grades"][n]["g"])).alignment = Alignment(horizontal='center', vertical='center')
                    except:
                        sheet1.cell(row=items+2,column=cs+3,value=res[items]["grades"][n]["g"]).alignment = Alignment(horizontal='center', vertical='center')
    sheet1.merge_cells(start_row=len(res)+2, start_column=1, end_row=len(res)+5, end_column=6)
    sheet1.cell(row=len(res)+2,column=1,value="1.表中数据来源须该班同学使用“西院助手”小程序访问并刷新该学期成绩\n2.留空为该同学还未刷新到最新，未使用小程序不会显示该同学行\n3.该表成绩为教务系统获取成绩，真实有效").alignment = Alignment(horizontal='center', vertical='center')
    sheet1.merge_cells(start_row=len(res)+2, start_column=7, end_row=len(res)+5, end_column=10)
    sheet1.cell(row=len(res)+2,column=7,value="生成时间：%s" % time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))).alignment = Alignment(horizontal='center', vertical='center')
    excel.save(saveUrl)

def get_pinfo(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd")
        }
        res = requests.post(url=myconfig.otherapi+"/info/pinfo",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["loginbad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求登录，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if Students.objects.filter(studentId=int(xh)):
            stu = Students.objects.get(studentId=int(xh))
            refreshTimes = int(stu.refreshTimes)
            try:
                startTime = time.time()
                lgn = Login(base_url=base_url)
                lgn.login(xh, pswd)
                if lgn.runcode == 1:
                    cookies = lgn.cookies
                    person = GetInfo(base_url=base_url, cookies=cookies)
                    pinfo = person.get_pinfo()
                    if pinfo.get("idNumber")[-6:] == pswd:
                        return HttpResponse(json.dumps({'err':"新生或专升本同学请在教务系统(jwxt.xcc.edu.cn)完善信息并审核且修改密码后登陆小程序！"}, ensure_ascii=False),
                                            content_type="application/json,charset=utf-8")
                    if pinfo.get('err'):
                        if pinfo.get('err') == "Connect Timeout":
                            return mywarn("登录超时","",xh,pswd)
                        else:
                            return pinfo
                    JSESSIONID = requests.utils.dict_from_cookiejar(cookies)["JSESSIONID"]
                    route = requests.utils.dict_from_cookiejar(cookies)["route"]
                    refreshTimes += 1
                    updateTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    Students.objects.filter(studentId=int(xh)).update(JSESSIONID=JSESSIONID, route=route,
                                                                      refreshTimes=refreshTimes, updateTime=updateTime)
                    endTime = time.time()
                    spendTime = endTime - startTime
                    print('【%s】登录了' % pinfo["name"])
                    content = ('【%s】[%s]第%d次访问登录了，耗时%.2fs' % (
                        datetime.datetime.now().strftime('%H:%M:%S'), pinfo["name"], refreshTimes, spendTime))
                    writeLog(content)
                    filename = ('Pinfo')
                    newData(xh, filename, json.dumps(pinfo, ensure_ascii=False))
                    return HttpResponse(json.dumps(pinfo, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                elif lgn.runcode == 2:
                    content = ('【%s】[%s]在登录时学号或者密码错误！' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    return HttpResponse(json.dumps({'err':'学号或者密码错误'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    content = ('【%s】[%s]在登录时网络或其它错误！' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    return HttpResponse(json.dumps({'err':'网络或token问题，请返回重试'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
            except Exception as e:
                if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                    # return get_pinfo(request)
                    return HttpResponse(json.dumps({'err':"请重新刷新一下"}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    content = ('【%s】[%s]登录时出错' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    traceback.print_exc()
                    return mywarn("登录未知错误",str(e),xh,pswd)
        else:
            try:
                startTime = time.time()
                lgn = Login(base_url=base_url)
                lgn.login(xh, pswd)
                if lgn.runcode == 1:
                    cookies = lgn.cookies
                    person = GetInfo(base_url=base_url, cookies=cookies)
                    pinfo = person.get_pinfo()
                    if pinfo.get("idNumber")[-6:] == pswd:
                        return HttpResponse(json.dumps({'err':"新生或专升本同学请在教务系统(jwxt.xcc.edu.cn)完善信息并审核且修改密码后登陆小程序！"}, ensure_ascii=False),
                                            content_type="application/json,charset=utf-8")
                    if pinfo.get('err'):
                        if pinfo.get('err') == "Connect Timeout":
                            return mywarn("登录超时","",xh,pswd)
                        else:
                            return pinfo
                    JSESSIONID = requests.utils.dict_from_cookiejar(cookies)["JSESSIONID"]
                    route = requests.utils.dict_from_cookiejar(cookies)["route"]
                    updateTime = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    newstu = Students.create(int(pinfo["studentId"]), pinfo["name"], calSex(pinfo["idNumber"]), pinfo["collegeName"],
                                             pinfo["majorName"], pinfo["className"], pinfo["phoneNumber"],
                                             pinfo["birthDay"], pinfo["graduationSchool"], pinfo["domicile"],
                                             pinfo["email"], pinfo["national"], pinfo["idNumber"],
                                             JSESSIONID, route, updateTime)
                    newstu.save()
                    endTime = time.time()
                    spendTime = endTime - startTime
                    print('【%s】第一次登录' % pinfo["name"])
                    content = ('【%s】[%s]第一次登录，耗时%.2fs' % (
                        datetime.datetime.now().strftime('%H:%M:%S'), pinfo["name"], spendTime))
                    writeLog(content)
                    filename = ('Pinfo')
                    newData(xh, filename, json.dumps(pinfo, ensure_ascii=False))
                    return HttpResponse(json.dumps(pinfo, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                elif lgn.runcode == 2:
                    content = ('【%s】[%s]在第一次登录时学号或者密码错误！' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    return HttpResponse(json.dumps({'err':'学号或者密码错误'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    content = ('【%s】[%s]在第一次登录时网络或其它错误！' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    return HttpResponse(json.dumps({'err':'网络或token问题，请返回重试'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
            except Exception as e:
                # print(e)
                if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                    # return get_pinfo(request)
                    return HttpResponse(json.dumps({'err':"请重新刷新一下"}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    content = ('【%s】[%s]第一次登录时出错' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
                    writeLog(content)
                    if str(e) == "'NoneType' object has no attribute 'get'":
                        return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                            content_type="application/json,charset=utf-8")
                    traceback.print_exc()
                    return mywarn("登录未知错误",str(e),xh,pswd)
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def refresh_class(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd")
        }
        res = requests.post(url=myconfig.otherapi+"/info/refreshclass",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["loginbad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求登录，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if not Students.objects.filter(studentId=int(xh)):
            content = ('【%s】[%s]未登录更新班级信息' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            stu = Students.objects.get(studentId=int(xh))
        try:
            startTime = time.time()
            print('【%s】更新了班级信息' % stu.name)
            JSESSIONID = str(stu.JSESSIONID)
            route = str(stu.route)
            cookies_dict = {
                'JSESSIONID': JSESSIONID,
                'route': route
            }
            cookies = requests.utils.cookiejar_from_dict(cookies_dict)
            person = GetInfo(base_url=base_url, cookies=cookies)
            nowClass = person.get_now_class()
            try:
                if nowClass.get('err'):
                    if nowClass.get('err') == "Connect Timeout":
                        return mywarn("更新班级超时","",xh,pswd)
            except:
                pass
            if stu.className == nowClass:
                return HttpResponse(json.dumps({'err':"你的班级并未发生变化~"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
            Students.objects.filter(studentId=int(xh)).update(className=nowClass)
            endTime = time.time()
            spendTime = endTime - startTime
            content = ('【%s】[%s]更新了班级信息，耗时%.2fs' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name, spendTime))
            writeLog(content)
            return HttpResponse(json.dumps({'success':"你已成功变更到【"+ nowClass + "】!",'class':nowClass}, ensure_ascii=False), content_type="application/json,charset=utf-8")
        except Exception as e:
            content = ('【%s】[%s]更新班级信息出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
            writeLog(content)
            if str(e) == "'NoneType' object has no attribute 'get'":
                    return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
            if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                return refresh_class(request)
            if 'Expecting value' not in str(e):
                traceback.print_exc()
                return mywarn("更新班级错误",str(e),xh,pswd)
            sta = update_cookies(xh, pswd)
            person = GetInfo(base_url=base_url, cookies=sta)
            nowClass = person.get_now_class()
            if stu.className == nowClass:
                return HttpResponse(json.dumps({'err':"你的班级并未发生变化~"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
            Students.objects.filter(studentId=int(xh)).update(className=nowClass)
            return HttpResponse(json.dumps({'success':"你已成功变更到【"+ nowClass + "】!",'class':nowClass}, ensure_ascii=False), content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def get_message(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd")
        }
        res = requests.post(url=myconfig.otherapi+"/info/message",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["jwxtbad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法访问（可能是学校机房断电或断网所致），小程序暂时无法登录和更新，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if not Students.objects.filter(studentId=int(xh)):
            content = ('【%s】[%s]未登录访问消息' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            stu = Students.objects.get(studentId=int(xh))
        try:
            startTime = time.time()
            # print('【%s】查看了消息' % stu.name)
            JSESSIONID = str(stu.JSESSIONID)
            route = str(stu.route)
            cookies_dict = {
                'JSESSIONID': JSESSIONID,
                'route': route
            }
            cookies = requests.utils.cookiejar_from_dict(cookies_dict)
            person = GetInfo(base_url=base_url, cookies=cookies)
            message = person.get_message()
            endTime = time.time()
            spendTime = endTime - startTime
            # content = ('【%s】[%s]访问了消息，耗时%.2fs' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name, spendTime))
            # writeLog(content)
            return HttpResponse(json.dumps(message, ensure_ascii=False), content_type="application/json,charset=utf-8")
        except Exception as e:
            if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                # return get_message(request)
                return HttpResponse(json.dumps({'err':"请重新刷新一下"}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
            else:
                content = ('【%s】[%s]访问消息出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
                writeLog(content)
                if str(e) == 'Expecting value: line 1 column 1 (char 0)':
                    return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                if str(e) != 'Expecting value: line 6 column 1 (char 11)':
                    traceback.print_exc()
                    return mywarn("消息请求错误",str(e),xh,pswd)
                sta = update_cookies(xh, pswd)
                person = GetInfo(base_url=base_url, cookies=sta)
                message = person.get_message()
                return HttpResponse(json.dumps(message, ensure_ascii=False), content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")


def get_study(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd"),
            'refresh':request.POST.get("refresh")
        }
        res = requests.post(url=myconfig.otherapi+"/info/study",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["studybad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求学业，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
            refresh = request.POST.get("refresh")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if not Students.objects.filter(studentId=int(xh)):
            content = ('【%s】[%s]未登录访问学业情况' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            stu = Students.objects.get(studentId=int(xh))
        if refresh == "no":
            filename = ('Study')
            cache = cacheData(xh, filename)
            if cache is not None:
                # print('cache')
                print('【%s】查看了学业缓存' % stu.name)
                return HttpResponse(json.dumps(cache, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            else:
                pass
        try:
            startTime = time.time()
            print('【%s】查看了学业情况' % stu.name)
            JSESSIONID = str(stu.JSESSIONID)
            route = str(stu.route)
            cookies_dict = {
                'JSESSIONID': JSESSIONID,
                'route': route
            }
            cookies = requests.utils.cookiejar_from_dict(cookies_dict)
            person = GetInfo(base_url=base_url, cookies=cookies)
            study = person.get_study(xh)
            if study.get("err") == 'Connect Timeout':
                sta = update_cookies(xh, pswd)
                person = GetInfo(base_url=base_url, cookies=sta)
                study = person.get_study(xh)
                gpa = str(study["gpa"]) if str(study["gpa"]) !="" or str(study["gpa"]) is not None else "init"
                Students.objects.filter(studentId=int(xh)).update(gpa=gpa)
                filename = ('Study')
                newData(xh, filename, json.dumps(study, ensure_ascii=False))
                return HttpResponse(json.dumps(study, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            endTime = time.time()
            spendTime = endTime - startTime
            content = ('【%s】[%s]访问了学业情况，耗时%.2fs' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name, spendTime))
            writeLog(content)
            gpa = str(study["gpa"]) if str(study["gpa"]) !="" or str(study["gpa"]) is not None else "init"
            Students.objects.filter(studentId=int(xh)).update(gpa=gpa)
            filename = ('Study')
            newData(xh, filename, json.dumps(study, ensure_ascii=False))
            return HttpResponse(json.dumps(study, ensure_ascii=False), content_type="application/json,charset=utf-8")
        except Exception as e:
            if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                # return get_study(request)
                return HttpResponse(json.dumps({'err':'更新出现问题，请待教务系统修复'}, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            elif "list index out of range" in str(e) and int(xh[0:2]) >= int(myconfig.nGrade[2:4]):
                return HttpResponse(json.dumps({'err':'暂无学业信息'}, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            else:
                content = ('【%s】[%s]访问学业情况出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
                writeLog(content)
                if str(e) != 'list index out of range':
                    traceback.print_exc()
                    return mywarn("学业请求错误",str(e),xh,pswd)
                sta = update_cookies(xh, pswd)
                person = GetInfo(base_url=base_url, cookies=sta)
                study = person.get_study(xh)
                gpa = str(study["gpa"]) if str(study["gpa"]) !="" or str(study["gpa"]) is not None else "init"
                Students.objects.filter(studentId=int(xh)).update(gpa=gpa)
                filename = ('Study')
                newData(xh, filename, json.dumps(study, ensure_ascii=False))
                return HttpResponse(json.dumps(study, ensure_ascii=False), content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")


def get_grade(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd"),
            'year':request.POST.get("year"),
            'term':request.POST.get("term"),
            'refresh':request.POST.get("refresh")
        }
        res = requests.post(url=myconfig.otherapi,data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["gradebad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求成绩，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
            year = request.POST.get("year")
            term = request.POST.get("term")
            refresh = request.POST.get("refresh")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if not Students.objects.filter(studentId=int(xh)):
            content = ('【%s】[%s]未登录访问成绩' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            stu = Students.objects.get(studentId=int(xh))
        if refresh == "no":
            filename = ('Grades-%s%s' % (str(year), str(term)))
            cache = cacheData(xh, filename)
            if cache is not None:
                # print('cache')
                def isLast(ny,nt,y,t):
                    ny = (myconfig.nGrade)[0:4]
                    nt = (myconfig.nGrade)[4:5]
                    if str(year) == ny:
                        pass
                    else:
                        if int(nt)-1 == 0 and int(term)==2:
                            pass
                        else:
                            print('【%s】查看了%s-%s的成绩缓存' % (stu.name, year, term))
                            return HttpResponse(json.dumps(cache, ensure_ascii=False),
                                            content_type="application/json,charset=utf-8")
            else:
                pass
        try:
            startTime = time.time()
            print('【%s】查看了%s-%s的成绩' % (stu.name, year, term))
            JSESSIONID = str(stu.JSESSIONID)
            route = str(stu.route)
            cookies_dict = {
                'JSESSIONID': JSESSIONID,
                'route': route
            }
            cookies = requests.utils.cookiejar_from_dict(cookies_dict)
            person = GetInfo(base_url=base_url, cookies=cookies)
            grade = person.get_grade(year, term)
            if grade.get("err"):
                if grade.get("err") == "Connect Timeout":
                    update_cookies(xh, pswd)
                    return mywarn("成绩超时","",xh,pswd)
                elif grade.get("err") == "No Data":
                    if int(xh[0:2]) > int(myconfig.nGrade[2:4]):
                        return HttpResponse(json.dumps({'err':"当前你还没有任何成绩信息"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
                    else:
                        return HttpResponse(json.dumps({'err':"还没有" + year+"-"+term + "学期的成绩，点击顶栏也看看以前的吧~"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
                elif grade.get("err") == "Error Term":
                    return HttpResponse(json.dumps({'err':"网络问题，请重新访问请求课程"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
            Students.objects.filter(studentId=int(xh)).update(gpa = grade.get("gpa") if grade.get("gpa")!="" or grade.get("gpa") is not None else "init")
            endTime = time.time()
            spendTime = endTime - startTime
            content = ('【%s】[%s]访问了%s-%s的成绩，耗时%.2fs' % (
                datetime.datetime.now().strftime('%H:%M:%S'), stu.name, year, term, spendTime))
            writeLog(content)
            
            filename = ('Grades-%s%s' % (str(year), str(term)))
            newData(xh, filename, json.dumps(grade, ensure_ascii=False))
            # print('write')

            return HttpResponse(json.dumps(grade, ensure_ascii=False), content_type="application/json,charset=utf-8")
        except Exception as e:
            # print(e)
            if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                # return get_grade(request)
                return HttpResponse(json.dumps({'err':"请重新刷新一下"}, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            else:
                content = ('【%s】[%s]访问成绩出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
                writeLog(content)
                if str(e) == 'Expecting value: line 1 column 1 (char 0)':
                    return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                if str(e) != 'Expecting value: line 3 column 1 (char 4)':
                    traceback.print_exc()
                    return mywarn("成绩请求错误",str(e),xh,pswd)
                sta = update_cookies(xh, pswd)
                person = GetInfo(base_url=base_url, cookies=sta)
                grade = person.get_grade(year, term)
                if grade.get("gpa") == "" or grade.get("gpa") is None:
                    return HttpResponse(json.dumps({'err':'平均学分绩点获取失败，请重试~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                Students.objects.filter(studentId=int(xh)).update(gpa = grade.get("gpa"))
                filename = ('Grades-%s%s' % (str(year), str(term)))
                newData(xh, filename, json.dumps(grade, ensure_ascii=False))

                return HttpResponse(json.dumps(grade, ensure_ascii=False), content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

# def get_grade2(request):
#     myconfig = Config.objects.all().first()
#     if myconfig.apichange:
#         data = {
#             'xh':request.POST.get("xh"),
#             'pswd':request.POST.get("pswd"),
#             'year':request.POST.get("year"),
#             'term':request.POST.get("term"),
#             'refresh':request.POST.get("refresh")
#         }
#         res = requests.post(url=myconfig.otherapi+"/info/grade",data=data)
#         return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
#                             content_type="application/json,charset=utf-8")
#     if myconfig.maintenance:
#         return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
#                             content_type="application/json,charset=utf-8")
#     # if mpconfig["gradebad"]:
#     #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求成绩，请待学校修复！'}, ensure_ascii=False),
#     #                             content_type="application/json,charset=utf-8")
#     if request.method == 'POST':
#         if request.POST:
#             xh = request.POST.get("xh")
#             pswd = request.POST.get("pswd")
#             year = request.POST.get("year")
#             term = request.POST.get("term")
#             refresh = request.POST.get("refresh")
#         else:
#             return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
#                                 content_type="application/json,charset=utf-8")
#         if not Students.objects.filter(studentId=int(xh)):
#             content = ('【%s】[%s]未登录访问成绩' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
#             writeLog(content)
#             return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
#                                 content_type="application/json,charset=utf-8")
#         else:
#             stu = Students.objects.get(studentId=int(xh))
#         if refresh == "no":
#             filename = ('GradesN-%s%s' % (str(year), str(term)))
#             cache = cacheData(xh, filename)
#             if cache is not None:
#                 # print('cache')
#                 print('【%s】查看了%s-%s的成绩缓存' % (stu.name, year, term))
#                 return HttpResponse(json.dumps(cache, ensure_ascii=False),
#                                     content_type="application/json,charset=utf-8")
#             else:
#                 pass
#         try:
#             startTime = time.time()
#             print('【%s】查看了%s-%s的成绩' % (stu.name, year, term))
#             JSESSIONID = str(stu.JSESSIONID)
#             route = str(stu.route)
#             cookies_dict = {
#                 'JSESSIONID': JSESSIONID,
#                 'route': route
#             }
#             cookies = requests.utils.cookiejar_from_dict(cookies_dict)
#             person = GetInfo(base_url=base_url, cookies=cookies)
#             grade = person.get_grade2(year, term)
#             if grade.get("err") == "请求超时，鉴于教务系统特色，已帮你尝试重新登录，重试几次，还不行请麻烦你自行重新登录，或者在关于里面反馈！当然，也可能是教务系统挂了~":
#                 update_cookies(xh, pswd)
#                 return HttpResponse(json.dumps({'err':grade.get("err")}, ensure_ascii=False), content_type="application/json,charset=utf-8")
#             if grade.get("err") == "看起来你这学期好像还没有出成绩，点击顶栏也看看以前的吧~":
#                 return HttpResponse(json.dumps({'err':grade.get("err")}, ensure_ascii=False), content_type="application/json,charset=utf-8")
#             Students.objects.filter(studentId=int(xh)).update(gpa = grade.get("gpa"))
#             endTime = time.time()
#             spendTime = endTime - startTime
#             content = ('【%s】[%s]访问了%s-%s的成绩，耗时%.2fs' % (
#                 datetime.datetime.now().strftime('%H:%M:%S'), stu.name, year, term, spendTime))
#             writeLog(content)
            
#             filename = ('GradesN-%s%s' % (str(year), str(term)))
#             newData(xh, filename, json.dumps(grade, ensure_ascii=False))
#             # print('write')

#             return HttpResponse(json.dumps(grade, ensure_ascii=False), content_type="application/json,charset=utf-8")
#         except Exception as e:
#             # print(e)
#             if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
#                 # return get_grade2(request)
#                 return HttpResponse(json.dumps({'err':'更新出现问题，请待教务系统修复'}, ensure_ascii=False),
#                                     content_type="application/json,charset=utf-8")
#             else:
#                 content = ('【%s】[%s]访问成绩出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
#                 writeLog(content)
#                 if str(e) == 'Expecting value: line 1 column 1 (char 0)':
#                     return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
#                                         content_type="application/json,charset=utf-8")
#                 if str(e) != 'Expecting value: line 3 column 1 (char 4)':
#                     traceback.print_exc()
#                     return mywarn("成绩请求错误",str(e),xh,pswd)
#                 sta = update_cookies(xh, pswd)
#                 person = GetInfo(base_url=base_url, cookies=sta)
#                 grade = person.get_grade2(year, term)
#                 if grade.get("gpa") == "" or grade.get("gpa") is None:
#                     return HttpResponse(json.dumps({'err':'平均学分绩点获取失败，请重试~'}, ensure_ascii=False),
#                                         content_type="application/json,charset=utf-8")
#                 Students.objects.filter(studentId=int(xh)).update(gpa = grade.get("gpa"))
#                 filename = ('GradesN-%s%s' % (str(year), str(term)))
#                 newData(xh, filename, json.dumps(grade, ensure_ascii=False))

#                 return HttpResponse(json.dumps(grade, ensure_ascii=False), content_type="application/json,charset=utf-8")
#     else:
#         return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
#                             content_type="application/json,charset=utf-8")

def get_schedule(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'pswd':request.POST.get("pswd"),
            'year':request.POST.get("year"),
            'term':request.POST.get("term"),
            'refresh':request.POST.get("refresh")
        }
        res = requests.post(url=myconfig.otherapi+"/info/schedule",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if myconfig.maintenance:
        return HttpResponse(json.dumps({'err':'教务系统出错维护中，请静待教务系统恢复正常！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    # if mpconfig["schedulebad"]:
    #     return HttpResponse(json.dumps({'err':'当前教务系统无法请求课表，请待学校修复！'}, ensure_ascii=False),
    #                             content_type="application/json,charset=utf-8")
    if request.method == 'POST':
        if request.POST:
            xh = request.POST.get("xh")
            pswd = request.POST.get("pswd")
            year = request.POST.get("year")
            term = request.POST.get("term")
            refresh = request.POST.get("refresh")
        else:
            return HttpResponse(json.dumps({'err':'请提交正确的post数据'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        if not Students.objects.filter(studentId=int(xh)):
            content = ('【%s】[%s]未登录访问课程' % (datetime.datetime.now().strftime('%H:%M:%S'), xh))
            writeLog(content)
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            stu = Students.objects.get(studentId=int(xh))
        if refresh == "no":
            filename = ('Schedules-%s%s' % (str(year), str(term)))
            cache = cacheData(xh, filename)
            if cache is not None:
                # print('cache')
                print('【%s】查看了%s-%s的课表缓存' % (stu.name, year, term))
                return HttpResponse(json.dumps(cache, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            else:
                pass
        try:
            startTime = time.time()
            print('【%s】查看了%s-%s的课程' % (stu.name, year, term))
            JSESSIONID = str(stu.JSESSIONID)
            route = str(stu.route)
            cookies_dict = {
                'JSESSIONID': JSESSIONID,
                'route': route
            }
            cookies = requests.utils.cookiejar_from_dict(cookies_dict)
            person = GetInfo(base_url=base_url, cookies=cookies)
            schedule = person.get_schedule(year, term)
            if schedule.get('err'):
                if schedule.get('err') == "Connect Timeout":
                    return mywarn("更新课程超时","",xh,pswd)
                elif schedule.get('err') == "Error Term":
                    return HttpResponse(json.dumps({'err':"网络问题，请重新访问请求课程"}, ensure_ascii=False), content_type="application/json,charset=utf-8")
            endTime = time.time()
            spendTime = endTime - startTime
            content = ('【%s】[%s]访问了%s-%s的课程，耗时%.2fs' % (
                datetime.datetime.now().strftime('%H:%M:%S'), stu.name, year, term, spendTime))
            writeLog(content)

            filename = ('Schedules-%s%s' % (str(year), str(term)))
            newData(xh, filename, json.dumps(schedule, ensure_ascii=False))
            # print('write')

            return HttpResponse(json.dumps(schedule, ensure_ascii=False), content_type="application/json,charset=utf-8")
        except Exception as e:
            if "Connection broken" in str(e) or 'ECONNRESET' in str(e):
                # return get_schedule(request)
                return HttpResponse(json.dumps({'err':"请重新刷新一下"}, ensure_ascii=False),
                                    content_type="application/json,charset=utf-8")
            else:
                content = ('【%s】[%s]访问课程出错' % (datetime.datetime.now().strftime('%H:%M:%S'), stu.name))
                writeLog(content)
                if str(e) == 'Expecting value: line 1 column 1 (char 0)':
                    return HttpResponse(json.dumps({'err':'教务系统挂掉了，请等待修复后重试~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                if str(e) != 'Expecting value: line 3 column 1 (char 4)':
                    traceback.print_exc()
                    return mywarn("课程请求错误",str(e),xh,pswd)
                sta = update_cookies(xh, pswd)
                person = GetInfo(base_url=base_url, cookies=sta)
                schedule = person.get_schedule(year, term)
                
                filename = ('Schedules-%s%s' % (str(year), str(term)))
                newData(xh, filename, json.dumps(schedule, ensure_ascii=False))

                return HttpResponse(json.dumps(schedule, ensure_ascii=False), content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({'err':'请使用post并提交正确数据'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def joinDetail(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        res = requests.get(url=myconfig.otherapi+"/info/joindetail?type=" + request.GET.get("type"))
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    type = request.GET.get("type")
    allUsers = Students.objects.filter().all().count()
    if type == 'college':
        detail = [{
            'collegeName': i["collegeName"],
            'collegeNum': Students.objects.filter(collegeName=i["collegeName"]).count()
        } for i in Students.objects.values('collegeName').distinct().order_by('collegeName')]
        ndetail = sorted(detail,key=lambda keys:keys['collegeNum'], reverse=True)
        res = {
            'allUsers': allUsers,
            'collegeNum': int(Students.objects.values('collegeName').distinct().order_by('collegeName').count()),
            'detail': ndetail
        }
    elif type == 'major':
        detail = [{
            'majorName': i["majorName"],
            'majorNum': Students.objects.filter(majorName=i["majorName"]).count()
        } for i in Students.objects.values('majorName').distinct().order_by('majorName')]
        ndetail = sorted(detail,key=lambda keys:keys['majorNum'], reverse=True)
        res = {
            'allUsers': allUsers,
            'majorNum': int(Students.objects.values('majorName').distinct().order_by('majorName').count()),
            'detail': ndetail
        }
    elif type == 'class':
        detail = [{
            'className': i["className"],
            'classNum': Students.objects.filter(className=i["className"]).count()
        } for i in Students.objects.values('className').distinct().order_by('className')]
        ndetail = sorted(detail,key=lambda keys:keys['classNum'], reverse=True)
        res = {
            'allUsers': allUsers,
            'classNum': int(Students.objects.values('className').distinct().order_by('className').count()),
            'detail': ndetail
        }
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")

def get_position(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        res = requests.get(url=myconfig.otherapi+"/info/position?xh=" + request.GET.get("xh"))
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    #print(request)
    xh = request.GET.get("xh")
    if xh is None:
        return HttpResponse(json.dumps({'err':'参数不全'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    if not Students.objects.filter(studentId=int(xh)):
        return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    else:
        stu = Students.objects.get(studentId=int(xh))
        majorName = stu.majorName
        className = stu.className
        majorNum = Students.objects.filter(majorName=majorName,studentId__startswith=int(xh[0:2])).all().count()
        classNum = Students.objects.filter(className=className).all().count()
        if stu.gpa == "init":
            gpa = "init"
            return HttpResponse(json.dumps({'gpa': gpa,'majorCount':0,'classCount':0,'majorNum':majorNum,'classNum':classNum,'nMajorCount':"init",'nClassCount':"init"}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            gpa = float(stu.gpa)
        majorCount = 1
        classCount = 1
        nMajorCount = 0
        nClassCount = 0
        for m in Students.objects.filter(majorName=majorName).all().order_by('-gpa'):
            if m.gpa == "init" and str(m.studentId)[0:2] == xh[0:2]:
                nMajorCount += 1
            elif m.gpa == "init" or str(m.studentId)[0:2] != xh[0:2]:
                pass
            elif gpa >= float(m.gpa):
                break
            else:
                majorCount += 1
        for c in Students.objects.filter(className=className).all().order_by('-gpa'):
            if c.gpa == "init":
                nClassCount += 1
            elif gpa >= float(c.gpa):
                break
            else:
                classCount += 1
        return HttpResponse(json.dumps({'gpa': gpa,'majorCount':majorCount,'nMajorCount':nMajorCount,'nClassCount':nClassCount,'classCount':classCount,'majorNum':majorNum,'classNum':classNum}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def searchTeacher(request):
    myconfig = Config.objects.all().first()
    if request.method == "GET":
        xh = request.GET.get("xh")
        tname = request.GET.get("tname")
        if myconfig.apichange:
            res = requests.get(url=myconfig.otherapi+"/info/steacher?xh=" + request.GET.get("xh") + "&tname=" + request.GET.get("tname"))
            return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
    elif request.method == "POST":
        xh = request.POST.get("xh")
        tname = request.POST.get("tname")
        if myconfig.apichange:
            data = {
                'xh':request.POST.get("xh"),
                'tname':request.POST.get("tname")
            }
            res = requests.post(url=myconfig.otherapi+"/info/steacher",data=data)
            return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        
    if xh is None or tname is None:
        return HttpResponse(json.dumps({'err': '参数不全'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    else:
        if not Students.objects.filter(studentId=int(xh)):
            return HttpResponse(json.dumps({'err':'还未登录，请重新登录！'}, ensure_ascii=False),
                                content_type="application/json,charset=utf-8")
        else:
            date = datetime.datetime.now().strftime('%Y-%m-%d')
            stu = Students.objects.filter(studentId=int(xh))
            thisStu = Students.objects.get(studentId=int(xh))
            lastTime = thisStu.searchTimes.split(',')[0]
            remainTimes = thisStu.searchTimes.split(',')[1]
            if lastTime == date:
                if remainTimes != '0':
                    searchList = []
                    for s in Teachers.objects.filter(name__contains=tname).order_by('name'):
                        item = {
                            'name': s.name,
                            'collegeName': s.collegeName,
                            'title': s.title,
                            'phoneNumber': s.phoneNumber
                        }
                        searchList.append(item)
                        content = ('【%s】%s学号查询[%s]' % (datetime.datetime.now().strftime('%H:%M:%S'), xh, tname))
                        writeLog(content)
                    if len(searchList) != 0:
                        nremainTimes = int(remainTimes) - 1
                        stu.update(searchTimes=lastTime+','+str(nremainTimes))
                    else:
                        nremainTimes = int(remainTimes)
                    return HttpResponse(json.dumps({'count': len(searchList),'result':searchList,'times':nremainTimes}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    return HttpResponse(json.dumps({'err': '同学，你今天的查询次数已满哦~'}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
            else:
                if thisStu.classMonitor == 1:
                    nlastTime = date
                    nremainTimes = '4'
                    ncontent = nlastTime + ',' + nremainTimes
                    stu.update(searchTimes=ncontent)
                    searchList = []
                    for s in Teachers.objects.filter(name__contains=tname).order_by('name'):
                        item = {
                            'name': s.name,
                            'collegeName': s.collegeName,
                            'title': s.title,
                            'phoneNumber': s.phoneNumber
                        }
                        searchList.append(item)
                        content = ('【%s】%s学号查询[%s]' % (datetime.datetime.now().strftime('%H:%M:%S'), xh, tname))
                        writeLog(content)
                    return HttpResponse(json.dumps({'count': len(searchList),'result':searchList,'times':int(nremainTimes)}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")
                else:
                    nlastTime = date
                    nremainTimes = '2'
                    ncontent = nlastTime + ',' + nremainTimes
                    stu.update(searchTimes=ncontent)
                    searchList = []
                    for s in Teachers.objects.filter(name__contains=tname).order_by('name'):
                        item = {
                            'name': s.name,
                            'collegeName': s.collegeName,
                            'title': s.title,
                            'phoneNumber': s.phoneNumber
                        }
                        searchList.append(item)
                        content = ('【%s】%s学号查询[%s]' % (datetime.datetime.now().strftime('%H:%M:%S'), xh, tname))
                        writeLog(content)
                    return HttpResponse(json.dumps({'count': len(searchList),'result':searchList,'times':int(nremainTimes)}, ensure_ascii=False),
                                        content_type="application/json,charset=utf-8")

def searchExcept(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        data = {
            'xh':request.POST.get("xh"),
            'tname':request.POST.get("tname"),
            'collegeName':request.POST.get("collegeName"),
            'content':request.POST.get("content")
        }
        res = requests.post(url=myconfig.otherapi+"/info/scallback",data=data)
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    xh = request.POST.get("xh")
    tname = request.POST.get("tname")
    collegeName = request.POST.get("college")
    content = request.POST.get("content")
    ServerChan = config["ServerChan"]
    text = "黄页反馈"
    if ServerChan == "none":
        return HttpResponse(json.dumps({'err':'反馈失败，管理员未打开反馈接口'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    else:
        requests.get(ServerChan + 'text=' + text + '&desp=' + str(xh) + '\n' + str(tname) + str(collegeName) + '\n' + str(content))
        return HttpResponse(json.dumps({'msg':'反馈成功'}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def classGrades(request):
    myconfig = Config.objects.all().first()
    if myconfig.apichange:
        res = requests.get(url=myconfig.otherapi+"/info/classgrades?className=" + request.GET.get("className") + "&yt=" + request.GET.get("yt"))
        return HttpResponse(json.dumps(json.loads(res.text), ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    className = request.GET.get("className")
    yt = request.GET.get("yt")
    year = yt[0:4]
    term = yt[4:5]
    studentIdList = []
    for i in Students.objects.filter(className=className).order_by("studentId"):
        studentIdList.append(i.studentId)
    res = []
    lastCourses = []
    try:
        lastStu = Students.objects.filter(className=className).order_by("-updateTime")[0].studentId
        with open('data/' + str(lastStu)[0:2] + '/' + str(lastStu) + '/Grades-' + yt + '.json') as l:
            lastReq = json.loads(l.read())
            for course in lastReq.get("course"):
                if course.get("courseNature") != "通识教育任选" and course.get("courseNature") != "无" and course.get("gradeNature") == "正常考试":
                    lastCourses.append(course.get("courseTitle"))
    except:
        lastStu = Students.objects.filter(className=className).order_by("-updateTime")[1].studentId
        with open('data/' + str(lastStu)[0:2] + '/' + str(lastStu) + '/Grades-' + yt + '.json') as l:
            lastReq = json.loads(l.read())
            for course in lastReq.get("course"):
                if course.get("courseNature") != "通识教育任选" and course.get("courseNature") != "无" and course.get("gradeNature") == "正常考试":
                    lastCourses.append(course.get("courseTitle"))
    for stu in studentIdList:
        nowUrl = 'data/' + str(stu)[0:2] + '/' + str(stu) + '/Grades-' + yt + '.json'
        try:
            with open(nowUrl,mode='r',encoding='UTF-8') as f:
                stuReq = json.loads(f.read())
                stuRes = {
                    'name':stuReq.get("name"),
                    'xh':stuReq.get("studentId"),
                    'grades':[{
                        'n':item.get("courseTitle"),
                        'g':item.get("grade")
                    }for item in stuReq["course"] if item.get("courseNature") != "通识教育任选" and item.get("courseNature") != "无" and item.get("gradeNature") == "正常考试"]
                }
                res.append(stuRes)
        except:
            res.append({'name':Students.objects.get(studentId=int(str(stu))).name,'xh':str(stu),'grades':[]})
    result = {'lastCourses':lastCourses,'res':res}
    writeToExcel(result,'data/classes/'+className+'.xlsx')
    try:
        file = open('data/classes/'+className+'.xlsx', 'rb')
    except:
        return HttpResponse(json.dumps({'error': "文件不存在"}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response["Content-Disposition"] = "attachment; filename*=UTF-8''{}".format(escape_uri_path(className)+'.xlsx')
    return response

def book_search(request):
    type = request.GET.get("type")
    content = request.GET.get("content")
    page = request.GET.get("page")
    result = Search()
    res = result.search_book(type,content,page)
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def book_detail(request):
    marc = request.GET.get("marc")
    result = Search()
    res = result.book_detail(marc)
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def library_info(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    lgn = PLogin()
    cookies = lgn.login(xh,ppswd)
    person = Personal(cookies)
    res = person.get_info()
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def library_list(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    lgn = PLogin()
    cookies = lgn.login(xh,ppswd)
    person = Personal(cookies)
    res = person.book_list()
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def library_hist(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    lgn = PLogin()
    cookies = lgn.login(xh,ppswd)
    person = Personal(cookies)
    res = person.book_hist()
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def library_paylist(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    lgn = PLogin()
    cookies = lgn.login(xh,ppswd)
    person = Personal(cookies)
    res = person.paylist()
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def library_paydetail(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    lgn = PLogin()
    cookies = lgn.login(xh,ppswd)
    person = Personal(cookies)
    res = person.paydetail()
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def school_card(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    page = request.POST.get("page")
    lgn = PLogin()
    cookies = lgn.plogin(xh,ppswd)
    person = Infos(cookies)
    res = person.school_card(page)
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def financial(request):
    xh = request.POST.get("xh")
    ppswd = request.POST.get("ppswd")
    page = request.POST.get("page")
    lgn = PLogin()
    cookies = lgn.plogin(xh,ppswd)
    person = Infos(cookies)
    res = person.financial(page)
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")

def award(request):
    if request.method == "POST":
        keyword = request.POST.get("keyword")
    else:
        keyword = request.GET.get("keyword")
    url = "http://xcctw.cn/app/index.php?keyword=" + keyword + "&i=2&c=entry&a=site&do=fm&m=yoby_cha&rid=13"
    res = requests.get(url=url)
    soup = BeautifulSoup(res.text,'lxml')
    if soup.find(class_="weui-msgbox"):
        return HttpResponse(json.dumps({'err':"没有查询到结果"}, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")
    list = soup.find_all(class_="weui-cell__bd")
    result = []
    for items in list:
        name = (items.find_all(class_="f16")[0].get_text()[3:]).strip()
        studentId = (items.find_all(class_="f16")[1].get_text()[3:]).strip()
        college = (items.find_all(class_="f16")[2].get_text()[5:]).strip()
        major = (items.find_all(class_="f16")[3].get_text()[3:]).strip()
        detail = (items.find_all(class_="f16")[4].get_text()[5:]).strip()
        number = (items.find_all(class_="f16")[5].get_text()[5:]).strip()
        items = {'name':name,'studentId':studentId,'college':college,'major':major,'detail':detail,'number':number}
        result.append(items)
    return HttpResponse(json.dumps(result, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")

def get_maps(request):
    if request.method == "GET":
        xh = request.GET.get("xh")
    elif request.method == "POST":
        xh = request.POST.get("xh")
    allIn = Students.objects.all().count()
    thisStu = Students.objects.get(studentId=int(xh))
    thisStuBirthDayAndMonth = (thisStu.birthDay)[5:]
    names = Students.objects.filter(name=thisStu.name).count() - 1
    birthDay = Students.objects.filter(birthDay=thisStu.birthDay).count() - 1
    birthDayAndMonth = Students.objects.filter(birthDay__contains=thisStuBirthDayAndMonth).count() - 1
    classBirthDay = Students.objects.filter(className=thisStu.className,birthDay=thisStu.birthDay).count() - 1
    classBirthDayAndMonth = Students.objects.filter(className=thisStu.className,birthDay__contains=thisStuBirthDayAndMonth).count() - 1
    graduationSchool = Students.objects.filter(graduationSchool=thisStu.graduationSchool).count() - 1
    classGraduationSchool = Students.objects.filter(className=thisStu.className,graduationSchool=thisStu.graduationSchool).count() - 1
    domicile = Students.objects.filter(domicile=thisStu.domicile).count() - 1
    classDomicile = Students.objects.filter(className=thisStu.className,domicile=thisStu.domicile).count() - 1
    res = {
        'allIn': allIn,
        'name': names,
        'birthDay': birthDay,
        'birthDayAndMonth': birthDayAndMonth,
        'classBirthDay': classBirthDay,
        'classBirthDayAndMonth': classBirthDayAndMonth,
        'graduationSchool': graduationSchool,
        'classGraduationSchool': classGraduationSchool,
        'domicile': domicile,
        'places':thisStu.domicile,
        'classDomicile': classDomicile
    }
    return HttpResponse(json.dumps(res, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")

def isMonitor(request):
    xh = request.GET.get("xh")
    if Students.objects.filter(studentId=int(xh)):
        thisStu = Students.objects.get(studentId=int(xh))
        res = {"code":200,"monitor":True if thisStu.classMonitor == 1 else False}
        return HttpResponse(json.dumps(res, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({"err":"没有这个同学"}, ensure_ascii=False),
                        content_type="application/json,charset=utf-8")

def freetime(request):
    myconfig = Config.objects.all().first()
    xh = request.GET.get("xh")
    term = request.GET.get("term") if request.GET.get("term") is not None else myconfig.nSchedule
    weeks = request.GET.get("weeks") if request.GET.get("weeks") is not None else myconfig.nowweek
    mode = request.GET.get("mode") if request.GET.get("mode") is not None else "1"
    datafile = 'data/' + xh[0:2] + "/" + xh + "/" + "Schedules-" + term + ".json"
    fullSections = [1,2,3,4,5,6,7,8,9,10,11,12]
    if os.path.exists(datafile):
        with open(datafile,mode='r',encoding='UTF-8') as f:
            schedule_data = json.loads(f.read())
        res = {"Mon":[],"Tue":[],"Wed":[],"Thu":[],"Fri":[]}
        for item in schedule_data["normalCourse"]:
            if item["courseWeekday"] == "1" and int(weeks) in item["includeWeeks"]:
                res["Mon"].extend(item["includeSection"])
            elif item["courseWeekday"] == "2" and int(weeks) in item["includeWeeks"]:
                res["Tue"].extend(item["includeSection"])
            elif item["courseWeekday"] == "3" and int(weeks) in item["includeWeeks"]:
                res["Wed"].extend(item["includeSection"])
            elif item["courseWeekday"] == "4" and int(weeks) in item["includeWeeks"]:
                res["Thu"].extend(item["includeSection"])
            elif item["courseWeekday"] == "5" and int(weeks) in item["includeWeeks"]:
                res["Fri"].extend(item["includeSection"])
            else:
                pass
        if mode == "1":
            res["Mon"] = diffList(fullSections,res["Mon"])
            res["Tue"] = diffList(fullSections,res["Tue"])
            res["Wed"] = diffList(fullSections,res["Wed"])
            res["Thu"] = diffList(fullSections,res["Thu"])
            res["Fri"] = diffList(fullSections,res["Fri"])
        return HttpResponse(json.dumps(res, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")
    else:
        return HttpResponse(json.dumps({"err":"原因：1.该同学没有使用“西院助手”小程序。2.没有在小程序请求过该学期课程信息。3.还未到该学期"}, ensure_ascii=False),
                            content_type="application/json,charset=utf-8")